import datetime

import xlsConverter as xlc
import pandas as pd
import openpyxl

"""
This module is used to parse the original data from an Excel file
Firstly, the method get_periods is used to extract the periods from the Excel file
Then, the method parse_file is used to check if the file is a xls or xlsx file
Finally, the method convert_workbook_to_dataframe is used to convert the workbook to a dataframe
"""


def get_periods(file_path, is_xls):
    """
    This method is used to extract the periods from the Excel file
    If the file is a xls file, the method get_periods from xlsConverter is used
    If the file is a xlsx file, the argument sheetnames from openpyxl is used
    :param file_path: The path to the Excel file
    :param is_xls: A boolean to check if the file is a xls file
    :return: A list of strings representing the periods
    """
    if is_xls:
        periods = xlc.get_periods(file_path)
    else:
        workbook = openpyxl.load_workbook(file_path)
        periods = workbook.sheetnames

    return periods


def parse_file(file_path, is_xls, periods):
    """
    This method is used to parse the Excel file If the file is a xls file, the method convert_to_xlsx from
    xlsConverter is used If the file is a xlsx file, the argument load_workbook from openpyxl is used In both cases
    the "if else" gives an openpyxl workbook which is then converted to a dataframe using the method
    convert_workbook_to_dataframe
    :param file_path: The path to the Excel file
    :param is_xls: A boolean to check if the file is a xls file
    :param periods: A list of strings representing the periods
    :return: A structured pandas dataframe
    """
    if is_xls:
        workbook = xlc.convert_to_xlsx(file_path)

    else:
        workbook = openpyxl.load_workbook(file_path)

    return convert_workbook_to_dataframe(workbook, periods)


def convert_workbook_to_dataframe(workbook, periods):
    """
    This method converts a multiple worksheets openpyxl workbook into a single structured pandas dataframe.
    The dataframe will contain the following columns: Name, Period, Test, Competence, Total, Result
    To do so, the method will iterate over the worksheets and extract the relevant data
    To extract the relevant data, break points are used to determine the end of the data
    These break points are determined by the methods find_blank_index and find_max_col_index for the rows and columns
    The method also addresses the problem where two different tests have the same name
     by using the method update_test_name_counts. NOTE that openpyxl is 1-indexed
    :param workbook: The multiple worksheets openpyxl workbook
    :param periods: A list of strings representing the periods
    :return: The structured pandas dataframe
    """

    # The data list will contain the records that will be used to create the dataframe
    data = []

    # In order to address the problem where two different tests have the same name
    test_name_counts = {}

    name_sheet = workbook["Nom"]
    # The break_index_row is the index where the names stop it is common to all the periods as it is in the "Nom" sheet
    break_index_row, students = find_blank_index(name_sheet)

    for period in workbook.sheetnames:
        if period in periods:
            sheet = workbook[period]
            # Extract the tests names and make sure that they are unique
            test_names = update_test_name_counts(sheet, test_name_counts)
            break_index_col = 0

            if period == "Nom":
                # In this case there is only one column containing the names
                break_index_col += 2

            else:
                # Else we need to find the maximum column index
                break_index_col += find_max_col_index(sheet)

            for row in range(4, break_index_row):
                student_name = students[row - 4]

                for col in range(3, break_index_col):
                    # Takes the unique test name to prevent duplicates
                    test_name = test_names[col]

                    total = sheet.cell(2, col).value
                    competence = str(sheet.cell(3, col).value)
                    result = sheet.cell(row, col).value

                    record = [student_name, period, test_name, competence, total, result]
                    data.append(record)

    df = pd.DataFrame(data, columns=['Name', 'Period', 'Test', 'Competence', 'Total', 'Result'])

    return df


def update_test_name_counts(sheet, test_name_counts):
    """
    This method is used to update the test name counts and make sure that the test names are unique
    It takes the sheet and the actual test_name_counts dictionary as arguments.
    The method will return a dictionary containing the unique test names updated considering the new tests found
    :param sheet: The new sheet we want to extract the test names from
    :param test_name_counts: The actual test name counts dictionary
    :return: A list of unique test names for the sheet
    """
    unique_test_names = {}
    col_index = 3
    while col_index < sheet.max_column:
        test_name = get_test_name(sheet, col_index)
        if test_name not in test_name_counts:
            test_name_counts[test_name] = 1
            # First occurrence of the test keeps the original name
            unique_test_names[col_index] = test_name
        else:
            test_name_counts[test_name] += 1
            new_name = f"{test_name} - {test_name_counts[test_name]}"
            unique_test_names[col_index] = new_name
        col_index += 1
    return unique_test_names


def get_test_name(sheet, col):
    """
    This method is used to get the test name from the sheet.
    The test name will be returned as a string
    If the test name is a float, it means it corresponds to an Excel xls date
    If the test name is a datetime, it means it corresponds to an Excel xlsx date
    Else it is a string
    In case of dates the method will convert it to a string with the format %d-%m-%Y
    :param sheet: The sheet we want to extract the test name from
    :param col: The column index where the test name is located
    :return: A string representing the test name
    """
    test_name = sheet.cell(1, col).value

    # If it is .xls then the date is a float
    if isinstance(test_name, float):
        # It means that the value is an Excel date
        # 1/1/1900 is a standard for Excel
        date_format = datetime.datetime(1900, 1, 1) + datetime.timedelta(days=test_name - 2)
        return date_format.strftime('%d-%m-%Y')
    # If it is .xlsx then the date is datetime with %H:%M:%S
    elif isinstance(test_name, datetime.datetime):
        return test_name.strftime("%d-%m-%Y")
    return test_name


def find_blank_index(name_sheet):
    """
    This method is used to find the index where the names stop in the "Nom" sheet and to collect the student names.
    The method uses a loop to find the index where the names stop
    :param name_sheet: The "Nom" sheet
    :return: The index of last row containing names and a list of names
    """
    # The rows before this index doesn't contain names
    i = 4
    students = []
    while i < name_sheet.max_row:
        if name_sheet.cell(i, 2).value is None:
            break
        else:
            students.append(name_sheet.cell(i, 2).value)
            i += 1
    return i, students


def find_max_col_index(sheet):
    """
    This method is used to find the maximum column index in the sheet.
    The method uses a loop to find the maximum column index.
    :param sheet: The sheet we want to find the maximum column index from
    :return: The index of the last column containing tests data
    """
    i = 3
    while i < sheet.max_column:
        if sheet.cell(3, i).value is None:
            break
        else:
            i += 1
    return i
