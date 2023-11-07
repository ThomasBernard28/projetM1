from openpyxl import Workbook
import pandas as pd
import numpy as np


def get_a_dataframe_by_sheet(wb, sheet_name):
    df = pd.DataFrame(wb[sheet_name].values)
    return df


def get_student_results_from_one_sheet(wb, student_name, sheet_name):
    worksheet = get_a_dataframe_by_sheet(wb, sheet_name)
    student_index = find_student_index(wb["Nom"], student_name)
    if worksheet.loc[student_index, 0] == student_index - 2:
        results = worksheet.loc[student_index, 2:].values
        indexes = []
        for result in range(len(results)):
            index = "Test Number " + str(result)
            indexes.append(index)
        df = pd.DataFrame(
            {
                "Test Number": indexes,
                "Results": np.array(results)
            }
        )
        df.fillna(-1, inplace=True)
        df = df.sort_values(by='Test Number')
        df = df.sort_values(by='Results', ascending=False)
        return df


def find_student_index(names_sheet, student_name):
    i = 4
    while i < names_sheet.max_row and names_sheet.cell(i, 2).value is not None:
        if names_sheet.cell(i, 2).value == student_name:
            return i - 1
        else:
            i += 1
    raise Exception("The student with the name : " + student_name + "wasn't found.")
