import openpyxl
import xlrd
import pandas as pd


def convert_to_xlsx(file_path):
    # As openpyxl doesn't support .xls we will convert it to .xlsx using xlrd
    workbook_xls = xlrd.open_workbook(file_path)
    wb = openpyxl.Workbook()

    # I use the nameSheet "blank" index of the name_sheet as a break index because
    # the list of results stops after the last student.
    name_sheet = workbook_xls.sheet_by_name("Nom")
    break_index_row = find_blank_index(name_sheet)

    class_name = name_sheet.cell_value(0, 0)
    students = find_all_students(name_sheet)

    # We will also delete useless sheets and useless infos
    for sheet_xls in workbook_xls.sheets():
        # This solution might not be the best, but it is easily editable
        # Another solution would be to offer the possibility to the user to choose
        # which sheets he wants to load.
        if sheet_xls.name in ["Nom", "B1", "B2", "Noel", "B3", "B4", "Exam. Juin"]:
            ws = wb.create_sheet(title=sheet_xls.name)

            break_index_col = 0

            if sheet_xls.name != "Nom":
                break_index_col += find_total_or_blank_index(sheet_xls)

            else:
                break_index_col += 2

            for row in range(break_index_row):
                row_data = []
                for col in range(break_index_col):
                    if sheet_xls.cell_type(row, col) == xlrd.XL_CELL_NUMBER:
                        row_data.append(sheet_xls.cell_value(row, col))

                    elif sheet_xls.cell_type(row, col) == xlrd.XL_CELL_EMPTY and (col > 1 and row > 2):
                        row_data.append("NP")

                    else:
                        row_data.append(sheet_xls.cell_value(row, col))

                ws.append(row_data)

            # Adding students names to all sheets
            for student in students:
                ws['B' + str(student[0])] = student[1]

            # Adding class name to all sheets
            ws['A1'] = class_name
        else:
            pass

    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    return wb


def find_total_or_blank_index(sheet_xls):
    # We start in the second col to avoid the class columns
    i = 2
    while i < sheet_xls.ncols:
        if sheet_xls.cell_value(0, i) == "Total SSFL" or sheet_xls.cell_value(0, i) == "":
            return i
        else:
            i += 1


def find_blank_index(sheet_xls):
    i = 3
    while i < sheet_xls.nrows:
        if sheet_xls.cell_value(i, 1) == "":
            return i
        else:
            i += 1


def find_all_students(sheet_xls):
    i = 3
    students = []
    while i < sheet_xls.nrows and sheet_xls.cell_value(i, 1) != "":
        students.append((i + 1, sheet_xls.cell_value(i, 1)))
        i += 1
    return students


def get_a_dataframe_by_sheet(wb, sheet_name):
    df = pd.DataFrame(wb[sheet_name].values)
    return df
