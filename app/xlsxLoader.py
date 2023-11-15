from openpyxl import Workbook, load_workbook


def parse_file(file_path):
    """
    This method will parse a .xlsx file by loading it in an openpyxl workbook, parsing and storing the result in a new
    openpyxl workbook.
    :param file_path: The path to the .xlsx file
    :return: The parsed workbook
    """

    workbook = load_workbook(file_path)
    parsed_workbook = Workbook()

    name_sheet = workbook["Nom"]
    break_index_row, students = find_blank_index(name_sheet)
    class_name = name_sheet.cell(1, 1).value

    for sheet_name in workbook.sheetnames:
        if sheet_name in ["Nom", "B1", "B2", "Noel", "B3", "B4", "Exam. Juin"]:
            sheet = workbook[sheet_name]
            parsed_sheet = parsed_workbook.create_sheet(title=sheet_name)

            break_index_col = 0

            if sheet_name == "Nom":
                break_index_col += 2

            else:
                break_index_col += find_total_or_blank_index(sheet)

            for row in range(1, break_index_row):
                row_data = []
                for col in range(1, break_index_col):
                    row_data.append(sheet.cell(row, col).value)

                parsed_sheet.append(row_data)

            for student in students:
                # Adding in the B colum the student name and in the A column the index
                parsed_sheet['B' + str(student[0])], parsed_sheet['A' + str(student[0])] = student[1], student[0] - 3

            parsed_sheet['A1'] = class_name

        else:
            pass

    default_sheet = parsed_workbook["Sheet"]
    parsed_workbook.remove(default_sheet)

    return parsed_workbook


def find_blank_index(name_sheet):
    # The rows before this index doesn't contain names
    i = 4
    students = []
    while i < name_sheet.max_row:
        if name_sheet.cell(i, 2).value is None:
            break
        else:
            students.append((i, name_sheet.cell(i, 2).value))
            i += 1
    return i, students


def find_total_or_blank_index(sheet):
    i = 3
    while i < sheet.max_column:
        if sheet.cell(1, i).value == "Total SSFL" or sheet.cell(1, i).value is None:
            break
        else:
            i += 1
    return i
